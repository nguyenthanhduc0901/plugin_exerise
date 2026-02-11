from collections.abc import Generator
from typing import Any
from io import BytesIO
import openpyxl
import xlrd

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage
from dify_plugin.file.file import File

# Import logging and custom handler
import logging
from dify_plugin.config.logger_format import plugin_logger_handler

# Set up logging with the custom handler
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
logger.addHandler(plugin_logger_handler)


NO_FILE_INPUT_PROVIDED_MSG = "ファイルがアップロードされていません。変更を反映するためにファイルをアップロードしてください。"
FILE_NOT_SUPPORTED_MSG = "アップロードされたファイル形式が正しくありません。Excel形式（.xlsx、.xls）のファイルをアップロードしてください。"
FILE_EMPTY_MSG = "アップロードされたファイルにデータが含まれていません。適切なデータが入力されたファイルをアップロードしてください。"
ACCEPTED_FILE_EXTENSIONS = [".xls", ".xlsx"]
# Standard header for Type 1 files
STANDARD_HEADER = [
    "検査種別(部位)", "御指摘内容", "立会", "顧客", "製品", "編成", "処置",
    "不良ｺｰﾄﾞ", "不良分類", "修正分類", "切粉寸法", "個数", "多数粉フラグ",
    "配電盤フラグ", "作業組", "実際の作業組", "運送情報", "",
    "要約", "カウント"
]

class Uc03FileCheck(Tool):
    
    """
    Tool to check if Excel files are valid (correct extension) and not empty.
    """

    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage]:
        """
        Invoke the file check tool.
        
        Parameters:
            tool_parameters: Dictionary containing tool parameters, specifically 'file_inputs'
        
        Returns:
            Generator yielding ToolInvokeMessage objects (text and variables)
        """
        file_inputs = tool_parameters.get("file_inputs")

        # 1. Check if any files were provided. If there are no files, Dify passes [None].
        if not file_inputs or file_inputs == [None]:
            logger.info("No file input provided.")
            yield self.create_text_message(NO_FILE_INPUT_PROVIDED_MSG)
            yield self.create_variable_message("success", "False")
            return

        # 2. Validate file extensions
        if not self._validate_file_extensions(file_inputs, ACCEPTED_FILE_EXTENSIONS):
            error_msg = f"File type not supported. Only {', '.join(ACCEPTED_FILE_EXTENSIONS)} are accepted."
            logger.info(error_msg)
            yield self.create_text_message(FILE_NOT_SUPPORTED_MSG)
            yield self.create_variable_message("success", "False")
            return

        # 3. Check if files are empty
        for file in file_inputs:
            if self._is_file_empty(file):
                error_msg = f"File '{file.filename}' is empty."
                logger.info(error_msg)
                yield self.create_text_message(FILE_EMPTY_MSG)
                yield self.create_variable_message("success", "False")
                return

        # If all checks pass
        logger.info("All files are valid and not empty.")
        yield self.create_variable_message("success", "True")
    
    def _validate_file_extensions(self, file_list: list[File], accepted_extensions: list[str]) -> bool:
        """
        Check if all files in the list have one of the accepted extensions.
        
        Parameters:
            file_list: List of File objects.
            accepted_extensions: List of allowed file extensions.
            
        Returns:
            True if all files are valid, False otherwise.
        """
        for file in file_list:
            file_extension = file.extension.lower()
            if not any(file_extension.endswith(ext) for ext in accepted_extensions):
                return False
        return True
    
    def _is_file_empty(self, file: File) -> bool:
        """
        Check if the given Excel file is empty (no data in any cell).
        
        Parameters:
            file: The File object to check.
            
        Returns:
            True if the file is empty, False otherwise.
        """
        file_extension = (file.extension or "").lower()
        blob = file.blob

        try:
            is_standard_header = False

            if file_extension == ".xlsx":
                workbook = openpyxl.load_workbook(
                    filename=BytesIO(blob),
                    read_only=True,
                    data_only=True,
                )
                
                if workbook.worksheets:
                    worksheet = workbook.worksheets[0]
                    
                    # Get first row as header
                    first_row = next(worksheet.iter_rows(values_only=True), None)
                    if first_row:
                        is_standard_header = self._is_standard_header(list(first_row))
                
                # Check if empty
                return self._check_xlsx_empty(blob, is_standard_header)

            elif file_extension == ".xls":
                book = xlrd.open_workbook(file_contents=blob)
                
                if book.nsheets > 0:
                    sheet = book.sheet_by_index(0)
                    
                    # Get first row as header
                    if sheet.nrows > 0:
                        first_row = [sheet.cell_value(0, col_index) for col_index in range(sheet.ncols)]
                        is_standard_header = self._is_standard_header(first_row)
                
                # Check if empty
                return self._check_xls_empty(blob, is_standard_header)

            return False
        except Exception as e:
            logger.error(f"Error checking file empty: {e}")
            return False

    def _is_standard_header(self, header_row: list) -> bool:
        """Check if the header row matches the standard header."""
        if not header_row or len(header_row) != len(STANDARD_HEADER):
            return False
        
        # Convert to string and strip whitespace for comparison
        header_str = [str(h).strip() if h is not None else "" for h in header_row]
        return header_str == STANDARD_HEADER

    def _check_xlsx_empty(self, blob: bytes, is_standard_header: bool) -> bool:
        """Check if XLSX file is empty based on file type."""
        workbook = openpyxl.load_workbook(
            filename=BytesIO(blob),
            read_only=True,
            data_only=True,
        )
        
        # Only check the first sheet
        worksheet = workbook.worksheets[0]
        
        # Determine starting row based on file type
        # Type 1 (standard header): skip row 1 (header), check from row 2
        # Type 2 (free format): check all rows starting from row 1
        start_row = 2 if is_standard_header else 1
        
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            # Skip rows before start_row
            if row_idx < start_row:
                continue
            
            for cell_value in row:
                if cell_value is None:
                    continue
                if isinstance(cell_value, str) and cell_value.strip() == "":
                    continue
                # Found data
                return False
        
        # No data found
        return True

    def _check_xls_empty(self, blob: bytes, is_standard_header: bool) -> bool:
        """Check if XLS file is empty based on file type."""
        book = xlrd.open_workbook(file_contents=blob)
        
        # Only check the first sheet
        sheet = book.sheet_by_index(0)
        
        # Determine starting row based on file type
        # Type 1 (standard header): skip row 1 (header), check from row 2
        # Type 2 (free format): check all rows starting from row 1
        start_row = 2 if is_standard_header else 1
        
        for row_index in range(sheet.nrows):
            # Skip rows before start_row (0-indexed, so start_row-1)
            if row_index < start_row - 1:
                continue
            
            for col_index in range(sheet.ncols):
                cell_value = sheet.cell_value(row_index, col_index)
                if cell_value is None:
                    continue
                if isinstance(cell_value, str) and cell_value.strip() == "":
                    continue
                # Found data
                return False
        
        # No data found
        return True


